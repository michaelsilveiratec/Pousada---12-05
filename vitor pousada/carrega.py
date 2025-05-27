from openpyxl import load_workbook

# Caminho do arquivo Excel
arquivo_excel = "reservas_hotel.xlsx"

# Carregar o arquivo Excel
workbook = load_workbook(arquivo_excel)

# Selecionar a planilha desejada (substitua 'cliente' pelo nome correto da planilha)
planilha = workbook["cliente"]

# Alterar dados específicos
for linha in range(2, planilha.max_row + 1):  # Começa na linha 2 para ignorar o cabeçalho
    nome = planilha.cell(row=linha, column=2).value  # Coluna 2: Nome
    if nome == "João Silva":  # Verifica o nome do cliente
        planilha.cell(row=linha, column=3).value = "novoemail@exemplo.com"  # Coluna 3: Email
        planilha.cell(row=linha, column=4).value = "99999-9999"  # Coluna 4: Telefone
        print(f"Dados alterados para o cliente '{nome}' na linha {linha}")

# Salvar as alterações no arquivo Excel
workbook.save(arquivo_excel)
print("Alterações salvas no arquivo Excel!")